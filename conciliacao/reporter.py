from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


STATUS_FILLS = {
    "Match Perfeito": PatternFill(start_color="D5F5E3", fill_type="solid"),
    "Glosa": PatternFill(start_color="FADBD8", fill_type="solid"),
    "Glosa + Pagamento a Maior": PatternFill(start_color="FADBD8", fill_type="solid"),
    "Nao Faturado": PatternFill(start_color="FADBD8", fill_type="solid"),
    "Pagamento a Maior": PatternFill(start_color="FEF9E7", fill_type="solid"),
    "Nao Encontrado": PatternFill(start_color="F6DDCC", fill_type="solid"),
}

HEADER_FILL = PatternFill(start_color="20515F", end_color="20515F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def to_excel(results: list, periodo_inicio: str, output_dir: str = ".") -> str:
    """Generate formatted Excel report from reconciliation results."""
    dt = datetime.strptime(periodo_inicio, "%d/%m/%Y")
    filename = f"{output_dir}/conciliacao_{dt.strftime('%Y-%m')}.xlsx"

    rows = []
    for r in results:
        rows.append({
            "Nome Faturamento": r["nome_pdf"] or "",
            "Nome Supabase": r["nome_supabase"] or "",
            "Score Match": r["score_match"] if r["score_match"] is not None else "",
            "Dias Esperados": len(r["datas_esperadas"]),
            "Dias Pagos": len(r["datas_pagas"]),
            "Datas Nao Pagas": ", ".join(r["datas_nao_pagas"]),
            "Datas Extras": ", ".join(r["datas_extras"]),
            "Status": r["status"],
        })

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Conciliacao", startrow=0)
        ws = writer.sheets["Conciliacao"]

        # Format header row
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Format data rows by status
        status_col_idx = 8  # Column H
        for row_idx in range(2, len(rows) + 2):
            status_value = ws.cell(row=row_idx, column=status_col_idx).value
            fill = STATUS_FILLS.get(status_value)
            if fill:
                for col_idx in range(1, 9):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        # Summary block
        summary_row = len(rows) + 3
        statuses = [r["status"] for r in results]
        total_glosa_dias = sum(len(r["datas_nao_pagas"]) for r in results)
        glosa_count = statuses.count("Glosa") + statuses.count("Glosa + Pagamento a Maior")

        summary = [
            ("Total de pacientes no PDF:", sum(1 for r in results if r["nome_pdf"])),
            ("Total de pacientes no Supabase:", sum(1 for r in results if r["nome_supabase"])),
            ("Match Perfeito:", statuses.count("Match Perfeito")),
            ("Glosas:", f"{glosa_count} ({total_glosa_dias} dias nao pagos)"),
            ("Nao Faturados:", statuses.count("Nao Faturado")),
            ("Nao Encontrados:", statuses.count("Nao Encontrado")),
        ]

        bold = Font(bold=True)
        for i, (label, value) in enumerate(summary):
            ws.cell(row=summary_row + i, column=1, value=label).font = bold
            ws.cell(row=summary_row + i, column=2, value=str(value))

    return filename
